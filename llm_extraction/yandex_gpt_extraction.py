import os
import json
import pandas as pd
import requests
import argparse
import re
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
import pathlib
from datetime import datetime

load_dotenv(override=True)

YANDEX_API_KEY = os.getenv("YANDEX_API_KEY")
YANDEX_FOLDER_ID = os.getenv("YANDEX_FOLDER_ID")
YANDEX_API_URL = "https://llm.api.cloud.yandex.net/foundationModels/v1/completion"

def get_structured_data(message, msg_id, abbreviations):
    """
    Извлекает структурированные данные из сообщения с помощью YandexGPT API
    """
    abbreviations_str = json.dumps(abbreviations, ensure_ascii=False)

    system_prompt = """
    Ты — эксперт по обработке сельскохозяйственных данных. Твоя задача — извлечь структурированные данные из текстовых сообщений о полевых работах, строго следуя предоставленным инструкциям и словарю сокращений. Возвращай данные только в указанном формате, игнорируя нерелевантную информацию. Если данные отсутствуют или не могут быть точно извлечены, возвращай пустой список JSON объектов.
    """

    user_prompt = f"""
    ### Задача
    Извлеки из сообщения информацию о полевых работах и представь её в виде списка JSON-объектов. Каждый объект представляет одну уникальную операцию на уровне Производственного Участка (ПУ) или, в виде исключения, Отделения (Отд).

    ### Словарь сокращений
    Используй следующий словарь для расшифровки аббревиатур операций и культур, а также для определения принадлежности ПУ и Отделений к Подразделениям:
    {abbreviations_str}

    ### Правила извлечения данных для каждого объекта:

    1.  **Идентификация Объекта:**
        * Каждый JSON-объект должен соответствовать **одной операции**.
        * Информация по операциям в сообщении всегда относится к одному Подразделению.
        * **Приоритет ПУ:** Если для операции есть данные 'По ПУ', объект создается на основе этих данных. Информация 'Отд' для той же операции игнорируется при создании объекта (используется только для определения подразделения, если нужно).
        * **Исключение Отд:** Если для операции данные 'По ПУ' **полностью отсутствуют**, объект создается на основе данных 'Отд [номер]'.

    2.  **Заполнение полей:** Для каждой идентифицированной операции заполни следующие поля:
        * **Дата**: дата строго в формате "ДД/ММ" или "ДД/ММ/ГГГГ", если указана, иначе "".
        * **Подразделение**:
            * **Всегда** указывай полное название подразделения ("АОР", "Восход", "ТСК", "Мир" и т.д.).
            * Определяй это название с помощью словаря сокращений, находя соответствие для номера ПУ или Отд, указанного в сообщении для данной операции.
            * **ВАЖНО:** Даже если данные по площади ('За день (га)', 'С начала операции (га)') берутся из строки 'Отд [номер]' (из-за отсутствия данных 'По ПУ'), в поле 'Подразделение' все равно должно стоять **название подразделения из словаря**, соответствующее этому 'Отд [номер]', а **не** сам текст "Отд [номер]".
        * **Операция**:
            * Полное название операции, расшифрованное с помощью словаря (например, "Пахота", "Дискование 2-е").
        * **Культура**:
            * Полное название культуры, расшифрованное с помощью словаря.
            * **ВАЖНОЕ ПРАВИЛО "ПОД":** Если в описании операции указано "[культура1] под [культура2]" (например, "сах св под пш"), то **ОСНОВНОЙ КУЛЬТУРОЙ СЧИТАЕТСЯ [культура2]** (в примере - "пш", т.е. "Пшеница озимая товарная"). Игнорируй [культуру1] ("сах св") при определении культуры для этого объекта.
            * Если указано просто "[операция] [культура]" (например, "Пахота соя"), используй эту культуру ("Соя товарная").
        * **За день (га)**:
            * Площадь за день в гектарах.
            * **ПРИОРИТЕТ ПУ:** **Всегда** извлекай это значение из строки `По ПУ [за день]/[с начала] га` или `По ПУ [за день] га/ с нарастающим [с начала] га`, если она есть.
            * **Исключение Отд:** Если строка 'По ПУ' для данной операции отсутствует, **только тогда** используй значение из строки `Отд [номер] [за день]/[с начала] га`.
        * **С начала операции (га)**:
            * Общая площадь с начала операции в гектарах.
            * **ПРИОРИТЕТ ПУ:** **Всегда** извлекай это значение из строки `По ПУ [за день]/[с начала] га` или `По ПУ [за день] га/ с нарастающим [с начала] га`, если она есть.
            * **Исключение Отд:** Если строка 'По ПУ' для данной операции отсутствует, **только тогда** используй значение из строки `Отд [номер] [за день]/[с начала] га`.
        * **Вал за день (ц)**:
            * Валовой сбор за день в центнерах (ц). Извлекай из "Вал [за день]/[с начала] кг".
            * **Перевод единиц:** Указанные в сообщении значения "Вал" даны в килограммах (кг). **ОБЯЗАТЕЛЬНО** переведи их в центнеры (ц), разделив на 100 (1 ц = 100 кг). Например, "Вал 1500/..." кг -> "Вал за день (ц)": "15".
            * Если данные отсутствуют, ставь "".
        * **Вал с начала (ц)**:
            * Общий валовой сбор с начала операции в центнерах (ц). Извлекай из "Вал [за день]/[с начала] кг".
            * **Перевод единиц:** Указанные в сообщении значения "Вал" даны в килограммах (кг). **ОБЯЗАТЕЛЬНО** переведи их в центнеры (ц), разделив на 100 (1 ц = 100 кг). Например, "Вал .../3000" кг -> "Вал с начала (ц)": "30".
            * Если данные отсутствуют, ставь "".

    ### Пример сообщения и ожидаемого вывода:

    **Сообщение:**

    Пахота зяби под сою
    По Пу 11/1451
    Отд 17 11/190

    Дисков сах св
    По Пу 118/1355
    Отд 17 118/653

    2-е диск сах св под пш
    По Пу 35/1177
    Отд 17 35/485

    Выравн зяби под сах св
    По Пу 51/1365
    Отд 11 51/461

    Выравн зяби под подс
    Отд 11 47/89

    **Ожидаемый вывод:**

    [
        {{
            "Дата": "",
            "Подразделение": "АОР",
            "Операция": "Пахота",
            "Культура": "Соя товарная"
            "За день (га)": "11"
            "С начала операции (га)": "1451"
            "Вал за день (ц)": "",
            "Вал с начала (ц)": ""
        }},
        {{
            "Дата": "",
            "Подразделение": "АОР", 
            "Операция": "Дискование",
            "Культура": "Свекла сахарная",
            "За день (га)": "118", 
            "С начала операции (га)": "1355", 
            "Вал за день (ц)": "",
            "Вал с начала (ц)": ""
        }},
        {{
            "Дата": "",
            "Подразделение": "АОР", 
            "Операция": "Дискование 2-е",
            "Культура": "Пшеница озимая товарная", 
            "За день (га)": "35", 
            "С начала операции (га)": "1177", 
            "Вал за день (ц)": "",
            "Вал с начала (ц)": ""
        }},
        {{
            "Дата": "",
            "Подразделение": "АОР", 
            "Операция": "Выравнивание зяби",
            "Культура": "Свекла сахарная", 
            "За день (га)": "51", 
            "С начала операции (га)": "1365", 
            "Вал за день (ц)": "",
            "Вал с начала (ц)": ""
        }},
        {{
            "Дата": "",
            "Подразделение": "АОР", 
            "Операция": "Выравнивание зяби",
            "Культура": "Подсолнечник товарный", 
            "За день (га)": "47", 
            "С начала операции (га)": "89", 
            "Вал за день (ц)": "",
            "Вал с начала (ц)": ""
        }}
    ]

    ### Обработай следующее сообщение:
    {message}
    """

    response_schema = {
        "type": "object",
        "properties": {
            "operations": {
                "type": "array",
                "items": {
                    "type": "object",
                    "properties": {
                        "Дата": {"type": "string"},
                        "Подразделение": {"type": "string"},
                        "Операция": {"type": "string"},
                        "Культура": {"type": "string"},
                        "За день (га)": {"type": "string"},
                        "С начала операции (га)": {"type": "string"},
                        "Вал за день (ц)": {"type": "string"},
                        "Вал с начала (ц)": {"type": "string"}
                    },
                    "required": [
                        "Дата", "Подразделение", "Операция", "Культура",
                        "За день (га)", "С начала операции (га)", "Вал за день (ц)", "Вал с начала (ц)"
                    ],
                    "additionalProperties": False
                }
            }
        },
        "required": ["operations"],
        "additionalProperties": False
    }

    try:
        headers = {
            "Authorization": f"Api-Key {YANDEX_API_KEY}",
            "Content-Type": "application/json"
        }

        data = {
            "modelUri": f"gpt://{YANDEX_FOLDER_ID}/yandexgpt-lite/rc",
            "completionOptions": {
                "temperature": 0.3,
                "maxTokens": 1500
            },
            "messages": [
                {
                    "role": "system",
                    "text": system_prompt
                },
                {
                    "role": "user",
                    "text": user_prompt
                }
            ],
            "response_format": {
                "type": "json_schema",
                "json_schema": {
                    "name": "structured_data",
                    "schema": response_schema,
                    "strict": True
                }
            }
        }

        response = requests.post(YANDEX_API_URL, headers=headers, json=data, timeout=30)
        response.raise_for_status()
        result = response.json()

        if "result" not in result or "alternatives" not in result["result"] or not result["result"]["alternatives"]:
            return "[]"

        text = result["result"]["alternatives"][0]["message"]["text"]
        if not text:
            return "[]"

        cleaned_text = text.strip()
        if cleaned_text.startswith("```") and cleaned_text.endswith("```"):
            cleaned_text = cleaned_text[3:-3].strip()
            
        cleaned_text = re.sub(r'//.*?(?=\n|$)', '', cleaned_text)

        structured_data = json.loads(cleaned_text)
        
        if isinstance(structured_data, dict):
            return json.dumps(structured_data.get("operations", []), ensure_ascii=False)
        else:
            return json.dumps(structured_data, ensure_ascii=False)

    except requests.exceptions.RequestException as e:
        print(f"Ошибка HTTP при обработке сообщения ID {msg_id}: {str(e)}")
        return "[]"
    except json.JSONDecodeError as e:
        print(f"Ошибка парсинга JSON для сообщения ID {msg_id}: {str(e)}")
        return "[]"
    except Exception as e:
        print(f"Ошибка при обработке сообщения ID {msg_id}: {str(e)}")
        return "[]"


def main():
    parser = argparse.ArgumentParser(description="Обработка chat_messages.json для извлечения данных о сельхозработах")
    parser.add_argument('--input', help='Путь к файлу chat_messages.json')
    parser.add_argument('--abbreviations', help='Путь к файлу с аббревиатурами')
    parser.add_argument('--output', help='Путь для сохранения Excel файла')
    args = parser.parse_args()

    script_path = pathlib.Path(__file__).resolve()
    script_dir = script_path.parent
    data_dir = script_dir.parent / "data"
    input_path = pathlib.Path(args.input) if args.input else data_dir / "chat_messages.json"
    abbreviations_path = pathlib.Path(args.abbreviations) if args.abbreviations else data_dir / "abbreviations.json"
    output_path = pathlib.Path(args.output) if args.output else data_dir / f"{datetime.now().strftime('%d-%m-%Y')}.xlsx"

    with open(abbreviations_path, 'r', encoding='utf-8') as file:
        abbreviations = json.load(file)

    with open(input_path, 'r', encoding='utf-8') as file:
        messages = json.load(file)

    print(f"Извлечено {len(messages)} сообщений")

    all_tables = []

    for i, msg in enumerate(messages[:2]):
        msg_id = i + 1
        msg_text = msg.get("text", "").strip('"') 
        msg_author = msg.get("author", "")
        msg_date = msg.get("date", "")

        print(f"Обработка сообщения {i+1}/{len(messages)} от {msg_author}")

        response = get_structured_data(msg_text, msg_id, abbreviations)

        try:
            structured_data = json.loads(response)
        except json.JSONDecodeError:
            structured_data = []  

        if structured_data:
            df = pd.DataFrame(structured_data)
            df['msg id'] = msg_id
            df['msg author'] = msg_author
            df['msg date'] = msg_date
            all_tables.append(df)
        else:
            empty_row = {
                'msg id': msg_id,
                'msg author': msg_author,
                'msg date': msg_date,
                'Дата': '',
                'Подразделение': '',
                'Операция': '',
                'Культура': '',
                'За день (га)': '',
                'С начала операции (га)': '',
                'Вал за день (ц)': '',
                'Вал с начала (ц)': ''
            }
            all_tables.append(pd.DataFrame([empty_row]))

    final_df = pd.concat(all_tables, ignore_index=True)

    cols = ["msg id", "msg author", "msg date", "Дата", "Подразделение", 
            "Операция", "Культура", "За день (га)", "С начала операции (га)", 
            "Вал за день (ц)", "Вал с начала (ц)"]

    final_df = final_df[cols]

    wb = Workbook()
    ws = wb.active
    ws.title = "Structured Messages"

    for r in dataframe_to_rows(final_df, index=False, header=True):
        ws.append(r)

    for row in range(2, len(final_df) + 2): 
        if all(cell.value is None or cell.value == '' for cell in ws[row][3:]): 
            for cell in ws[row]:
                cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    wb.save(output_path)

    print(f"Данные успешно экспортированы в '{output_path}'")
    print(f"Всего обработано {len(messages)} сообщений")
    print(f"Извлечено {len(final_df)} записей о сельхозработах")

if __name__ == "__main__":
    main()